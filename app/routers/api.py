"""API 路由：全部业务写操作 + 查询接口。HTMX 友好（写操作返回 HX-Redirect）。"""
import hashlib
import io
import json
import re
import zipfile
from datetime import date, datetime, timedelta

from fastapi import APIRouter, Request, UploadFile, File, Form, HTTPException
from fastapi.responses import Response, JSONResponse
from sqlalchemy import text as sqltext

from ..config import config
from ..database import SessionLocal
from ..models import (Contract, QuickReply, KnowledgeDoc, KnowledgeChunk,
                      Interview, TodoItem, AuditLog, Requisition, Candidate,
                      CandidateSection, Employee, EmployeeEvent,
                      AttendanceException, SalaryRecord, DocTemplate,
                      ChecklistItem)
from ..abilities.doc_utils import extract_text, split_articles
from ..abilities.resume_parser import parse_resume
from ..services.screening import screen, serialize

router = APIRouter(prefix="/api")

ARTICLE_RE = re.compile(r"第[一二三四五六七八九十百千零〇0-9]+[条款]")


# ---------- 工具 ----------

def _audit(db, action, target, detail=""):
    db.add(AuditLog(action=action, target=target, detail=detail[:1000]))


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    s = s.replace("/", "-").replace(".", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _redirect(url: str) -> Response:
    return Response(status_code=303, headers={"HX-Redirect": url, "Location": url})


def _rescan_contracts():
    """合同变更后重扫待办（单机量小，毫秒级同步执行）。"""
    try:
        from ..services.reminders import sync_contract_todos
        sync_contract_todos()
    except Exception:
        pass


def _rescan_interviews():
    try:
        from ..services.reminders import sync_interview_todos
        sync_interview_todos()
    except Exception:
        pass


@router.get("/health")
def health():
    return {"ok": True, "name": "人事工作台", "time": datetime.now().isoformat()}


@router.post("/login")
def login(password: str = Form(...)):
    if password == config.access_password and config.access_password:
        token = hashlib.sha256(password.encode("utf-8")).hexdigest()
        resp = _redirect("/")
        resp.set_cookie("hr_auth", token, httponly=True, samesite="lax")
        return resp
    return _redirect("/login?err=1")


# ---------- 合同 ----------

@router.post("/contract/add")
def contract_add(name: str = Form(...), personnel: str = Form(...),
                 start_date: str = Form(""), end_date: str = Form(""),
                 is_indefinite: str = Form(""), contract_no: int = Form(1),
                 contract_type: str = Form("固定期限"), status: str = Form("履行中"),
                 probation_start: str = Form(""), probation_end: str = Form(""),
                 note: str = Form("")):
    db = SessionLocal()
    try:
        c = Contract(
            name=name.strip(), personnel=personnel.strip(),
            start_date=_parse_date(start_date) if start_date else None,
            end_date=_parse_date(end_date) if end_date else None,
            is_indefinite=bool(is_indefinite) or contract_type == "无固定期限",
            contract_no=contract_no, contract_type=contract_type,
            status=status,
            probation_start=_parse_date(probation_start) if probation_start else None,
            probation_end=_parse_date(probation_end) if probation_end else None,
            note=note.strip())
        db.add(c)
        _audit(db, "add", "contract", f"{personnel}《{name}》")
        db.commit()
        _rescan_contracts()
        return _redirect("/contract")
    finally:
        db.close()


@router.post("/contract/{cid}/update")
def contract_update(cid: int, name: str = Form(...), personnel: str = Form(...),
                    start_date: str = Form(""), end_date: str = Form(""),
                    is_indefinite: str = Form(""), contract_no: int = Form(1),
                    contract_type: str = Form("固定期限"), status: str = Form("履行中"),
                    probation_start: str = Form(""), probation_end: str = Form(""),
                    note: str = Form("")):
    db = SessionLocal()
    try:
        c = db.get(Contract, cid)
        if not c:
            raise HTTPException(404, "合同不存在")
        c.name, c.personnel = name.strip(), personnel.strip()
        c.start_date = _parse_date(start_date) if start_date else None
        c.end_date = _parse_date(end_date) if end_date else None
        c.is_indefinite = bool(is_indefinite) or contract_type == "无固定期限"
        c.contract_no, c.contract_type, c.status = contract_no, contract_type, status
        c.probation_start = _parse_date(probation_start) if probation_start else None
        c.probation_end = _parse_date(probation_end) if probation_end else None
        c.note = note.strip()
        _audit(db, "update", "contract", f"id={cid}")
        db.commit()
        _rescan_contracts()
        return _redirect("/contract")
    finally:
        db.close()


@router.post("/contract/{cid}/delete")
def contract_delete(cid: int):
    db = SessionLocal()
    try:
        c = db.get(Contract, cid)
        if c:
            _audit(db, "delete", "contract", f"{c.personnel}《{c.name}》")
            # 清理该合同生成的待办，避免孤儿待办
            db.query(TodoItem).filter(TodoItem.kind == "contract",
                                      TodoItem.ref_id == cid,
                                      TodoItem.status == "open").delete()
            db.delete(c)
            db.commit()
        return _redirect("/contract")
    finally:
        db.close()


@router.get("/contract/template")
def contract_template():
    """下载 Excel 导入模板"""
    import io
    import openpyxl
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同台账"
    ws.append(["合同名称", "人员", "开始日期", "结束日期", "第几次合同",
               "合同类型", "状态", "试用期开始", "试用期结束", "备注"])
    ws.append(["劳动合同", "张三", "2024-09-01", "2027-08-31", "1",
               "固定期限", "履行中", "2024-09-01", "2024-11-30", ""])
    ws.append(["劳动合同", "李四", "2023-03-15", "2026-03-14", "2",
               "固定期限", "履行中", "", "", "第2次，到期前90天有特别提醒"])
    for col, w in zip("ABCDEFGHIJ", [14, 10, 12, 12, 12, 12, 10, 12, 12, 20]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contract_template.xlsx"})


@router.post("/contract/import")
async def contract_import(file: UploadFile = File(...)):
    import openpyxl
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, "Excel 为空")
        header = [str(h or "").strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(header)}
        need = ["合同名称", "人员"]
        for n in need:
            if n not in idx:
                raise HTTPException(400, f"缺少列：{n}（表头：{header}）")

        added = 0
        for r in rows[1:]:
            if not r or not any(str(x or "").strip() for x in r):
                continue
            name = str(r[idx["合同名称"]] or "").strip()
            personnel = str(r[idx["人员"]] or "").strip()
            if not name or not personnel:
                continue
            def col(key):
                return r[idx[key]] if key in idx else None
            ctype = str(col("合同类型") or "固定期限").strip() or "固定期限"
            db.add(Contract(
                name=name, personnel=personnel,
                start_date=_parse_date(col("开始日期")),
                end_date=_parse_date(col("结束日期")),
                is_indefinite=ctype == "无固定期限" or str(col("无固定期限") or "").strip() in ("是", "Y", "1"),
                contract_no=int(col("第几次合同") or 1) if str(col("第几次合同") or "").strip() else 1,
                contract_type=ctype,
                status=str(col("状态") or "履行中").strip() or "履行中",
                probation_start=_parse_date(col("试用期开始")),
                probation_end=_parse_date(col("试用期结束")),
                note=str(col("备注") or "").strip()))
            added += 1
        _audit(db, "import", "contract", f"文件{file.filename}，新增{added}条")
        db.commit()
        _rescan_contracts()
        return _redirect("/contract")
    finally:
        db.close()


# ---------- 常用回复 ----------

@router.post("/reply/add")
def reply_add(category: str = Form("其他"), title: str = Form(...),
              keywords: str = Form(""), content: str = Form(...),
              shortcut: str = Form("")):
    db = SessionLocal()
    try:
        sc = shortcut.strip().lower()
        if sc and db.query(QuickReply).filter(
                QuickReply.shortcut == sc, QuickReply.is_active.is_(True)).first():
            return _redirect("/replies?err=dup_shortcut")
        db.add(QuickReply(category=category, title=title.strip(),
                          keywords=keywords.strip(), content=content.strip(),
                          shortcut=sc))
        _audit(db, "add", "reply", title)
        db.commit()
        return _redirect("/replies")
    finally:
        db.close()


@router.post("/reply/{rid}/update")
def reply_update(rid: int, category: str = Form("其他"), title: str = Form(...),
                 keywords: str = Form(""), content: str = Form(...),
                 shortcut: str = Form("")):
    db = SessionLocal()
    try:
        r = db.get(QuickReply, rid)
        if not r:
            raise HTTPException(404, "回复不存在")
        sc = shortcut.strip().lower()
        if sc and db.query(QuickReply).filter(
                QuickReply.shortcut == sc, QuickReply.id != rid,
                QuickReply.is_active.is_(True)).first():
            return _redirect("/replies?err=dup_shortcut")
        r.category, r.title = category, title.strip()
        r.keywords, r.content = keywords.strip(), content.strip()
        r.shortcut = sc
        _audit(db, "update", "reply", title)
        db.commit()
        return _redirect("/replies")
    finally:
        db.close()


@router.post("/reply/{rid}/delete")
def reply_delete(rid: int):
    db = SessionLocal()
    try:
        r = db.get(QuickReply, rid)
        if r:
            db.delete(r)
            _audit(db, "delete", "reply", r.title)
            db.commit()
        return _redirect("/replies")
    finally:
        db.close()


@router.post("/reply/{rid}/copy")
def reply_copy(rid: int):
    """复制计数（点击复制时调用）"""
    db = SessionLocal()
    try:
        r = db.get(QuickReply, rid)
        if r:
            r.usage_count = (r.usage_count or 0) + 1
            db.commit()
        return Response("")
    finally:
        db.close()


# ---------- 知识库 ----------

@router.post("/knowledge/upload")
async def knowledge_upload(file: UploadFile = File(...),
                           title: str = Form(""), category: str = Form("制度"),
                           version: str = Form(""), effective_date: str = Form("")):
    raw = await file.read()
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(400, "文件超过 20MB")
    text = extract_text(file.filename, raw)
    if len(text.strip()) < 10:
        raise HTTPException(400, "未能从文档中提取到文本（可能是扫描件，一期需先转文字版）")

    db = SessionLocal()
    try:
        save_dir = config.data_dir / "knowledge"
        save_dir.mkdir(parents=True, exist_ok=True)
        fname = f"{datetime.now():%Y%m%d%H%M%S}_{file.filename}"
        dest = save_dir / fname
        dest.write_bytes(raw)

        doc = KnowledgeDoc(
            title=title.strip() or file.filename,
            category=category, version=version.strip(),
            effective_date=_parse_date(effective_date) if effective_date else None,
            source_file=file.filename, file_path=str(dest))
        db.add(doc)
        db.flush()

        chunks = split_articles(text)
        for seq, (st, body) in enumerate(chunks, start=1):
            kc = KnowledgeChunk(doc_id=doc.id, seq=seq, section_title=st, content=body)
            db.add(kc)
        doc.chunk_count = len(chunks)
        _audit(db, "add", "knowledge", f"{doc.title}，切片{len(chunks)}条")
        db.commit()
        return _redirect("/knowledge")
    finally:
        db.close()


@router.post("/knowledge/{did}/delete")
def knowledge_delete(did: int):
    db = SessionLocal()
    try:
        doc = db.get(KnowledgeDoc, did)
        if doc:
            db.query(KnowledgeChunk).filter(KnowledgeChunk.doc_id == did).delete()
            _audit(db, "delete", "knowledge", doc.title)
            db.delete(doc)
            db.commit()
        return _redirect("/knowledge")
    finally:
        db.close()


@router.post("/knowledge/search")
def knowledge_search(q: str = Form(...), limit: int = Form(10)):
    q = q.strip()
    if not q:
        return JSONResponse({"items": []})
    db = SessionLocal()
    try:
        terms = [t for t in re.split(r"[\s,，;；]+", q) if t][:5]
        if not terms:
            return JSONResponse({"items": []})
        # 检索统一用 LIKE 子串匹配：单机知识库切片量小（千级），全表扫描 <50ms 完全够用，
        # 无需引入 FTS/向量索引，避免维护成本与索引墓碑问题。
        where = " AND ".join(f"kc.content LIKE :p{i}" for i in range(len(terms)))
        params = {f"p{i}": f"%{t}%" for i, t in enumerate(terms)}
        params["lim"] = limit
        rows = db.execute(sqltext(
            f"SELECT kc.id, kd.title, kd.category, kd.version, kd.effective_date, "
            f"kc.section_title, kc.content "
            f"FROM knowledge_chunk kc JOIN knowledge_doc kd ON kd.id = kc.doc_id "
            f"WHERE {where} ORDER BY kc.id DESC LIMIT :lim"),
            params).fetchall()
        items = [{
            "title": r[1], "category": r[2], "version": r[3],
            "effective_date": str(r[4]) if r[4] else "",
            "section": r[5] or "", "content": r[6][:600],
        } for r in rows]
        return JSONResponse({"items": items})
    finally:
        db.close()


# ---------- 面试 ----------

@router.post("/interview/add")
def interview_add(candidate_name: str = Form(...), position: str = Form(""),
                  round_name: str = Form("初试"), interviewers: str = Form(""),
                  start_time: str = Form(""), end_time: str = Form(""),
                  mode: str = Form("现场"), location: str = Form(""),
                  status: str = Form("待确认"), note: str = Form("")):
    db = SessionLocal()
    try:
        def _dt(s):
            if not s:
                return None
            try:
                return datetime.strptime(s, "%Y-%m-%dT%H:%M")
            except ValueError:
                try:
                    return datetime.strptime(s, "%Y-%m-%d %H:%M")
                except ValueError:
                    return None
        db.add(Interview(
            candidate_name=candidate_name.strip(), position=position.strip(),
            round_name=round_name, interviewers=interviewers.strip(),
            start_time=_dt(start_time), end_time=_dt(end_time),
            mode=mode, location=location.strip(), status=status, note=note.strip()))
        _audit(db, "add", "interview", candidate_name)
        db.commit()
        _rescan_interviews()
        return _redirect("/interviews")
    finally:
        db.close()


@router.post("/interview/{iid}/update")
def interview_update(iid: int, candidate_name: str = Form(...), position: str = Form(""),
                     round_name: str = Form("初试"), interviewers: str = Form(""),
                     start_time: str = Form(""), end_time: str = Form(""),
                     mode: str = Form("现场"), location: str = Form(""),
                     status: str = Form("待确认"), note: str = Form("")):
    db = SessionLocal()
    try:
        iv = db.get(Interview, iid)
        if not iv:
            raise HTTPException(404, "面试不存在")
        def _dt(s):
            if not s:
                return None
            try:
                return datetime.strptime(s, "%Y-%m-%dT%H:%M")
            except ValueError:
                try:
                    return datetime.strptime(s, "%Y-%m-%d %H:%M")
                except ValueError:
                    return None
        iv.candidate_name = candidate_name.strip()
        iv.position, iv.round_name = position.strip(), round_name
        iv.interviewers, iv.start_time = interviewers.strip(), _dt(start_time)
        iv.end_time, iv.mode = _dt(end_time), mode
        iv.location, iv.status, iv.note = location.strip(), status, note.strip()
        _audit(db, "update", "interview", f"id={iid}")
        db.commit()
        _rescan_interviews()
        return _redirect("/interviews")
    finally:
        db.close()


@router.post("/interview/{iid}/status")
def interview_status(iid: int, status: str = Form(...)):
    db = SessionLocal()
    try:
        iv = db.get(Interview, iid)
        if iv:
            iv.status = status
            _audit(db, "status", "interview", f"id={iid} -> {status}")
            db.commit()
            _rescan_interviews()
        return _redirect("/interviews")
    finally:
        db.close()


@router.post("/interview/{iid}/delete")
def interview_delete(iid: int):
    db = SessionLocal()
    try:
        iv = db.get(Interview, iid)
        if iv:
            _audit(db, "delete", "interview", iv.candidate_name)
            db.query(TodoItem).filter(TodoItem.kind == "interview",
                                      TodoItem.ref_id == iid,
                                      TodoItem.status == "open").delete()
            db.delete(iv)
            db.commit()
        return _redirect("/interviews")
    finally:
        db.close()


# ---------- 待办 ----------

@router.post("/todo/add")
def todo_add(title: str = Form(...), due_date: str = Form(""),
             level: str = Form("medium"), kind: str = Form("system")):
    db = SessionLocal()
    try:
        db.add(TodoItem(title=title.strip(), kind=kind, due_date=_parse_date(due_date) if due_date else None,
                        level=level, status="open", source="手动添加"))
        _audit(db, "add", "todo", title)
        db.commit()
        return _redirect("/")
    finally:
        db.close()


@router.post("/todo/{tid}/done")
def todo_done(tid: int):
    db = SessionLocal()
    try:
        t = db.get(TodoItem, tid)
        if t:
            t.status = "done"
            t.resolved_at = datetime.now()
            _audit(db, "done", "todo", t.title)
            db.commit()
        return Response("")
    finally:
        db.close()


# ---------- 简历：岗位 ----------

@router.post("/requisition/add")
def requisition_add(title: str = Form(...), department: str = Form(""),
                    education: str = Form("不限"), min_years: int = Form(0),
                    city: str = Form("不限"), salary_max: int = Form(0),
                    skills: str = Form(""), skills_mode: str = Form("any")):
    db = SessionLocal()
    try:
        hard = {"education": education, "min_years": min_years, "city": city,
                "salary_max": salary_max, "skills": skills, "skills_mode": skills_mode}
        soft = {"weights": {"years": 20, "school": 12, "management": 14,
                            "stability": 12, "skills": 20},
                "pass_threshold": 80, "maybe_threshold": 60}
        rq = Requisition(title=title.strip(), department=department.strip(),
                         hard_json=json.dumps(hard, ensure_ascii=False),
                         soft_json=json.dumps(soft, ensure_ascii=False))
        db.add(rq)
        _audit(db, "add", "requisition", title)
        db.commit()
        return _redirect("/resume")
    finally:
        db.close()


@router.post("/requisition/{rid}/delete")
def requisition_delete(rid: int):
    db = SessionLocal()
    try:
        rq = db.get(Requisition, rid)
        if rq:
            _audit(db, "delete", "requisition", rq.title)
            db.delete(rq)
            db.commit()
        return _redirect("/resume")
    finally:
        db.close()


# ---------- 简历：导入 ----------

RESUME_EXTS = (".pdf", ".docx", ".txt", ".md")


def _save_and_parse(db, filename, raw, channel, requisition_id):
    fhash = hashlib.sha256(raw).hexdigest()
    if db.query(Candidate).filter(Candidate.file_hash == fhash).first():
        return "dup"
    try:
        result = parse_resume(filename, raw)
    except Exception:
        return "skip"   # 单个坏文件不中断整批
    save_dir = config.data_dir / "resumes"
    save_dir.mkdir(parents=True, exist_ok=True)
    fname = f"{datetime.now():%Y%m%d%H%M%S%f}_{filename}"
    dest = save_dir / fname
    dest.write_bytes(raw)

    f = result["fields"]
    try:
        skills_json = json.dumps(f.get("skills") or [], ensure_ascii=False)
    except Exception:
        skills_json = "[]"
    cand = Candidate(
        requisition_id=requisition_id,
        name=f.get("name", ""), phone=f.get("phone", ""),
        email=f.get("email", ""), city=f.get("city", ""),
        education=f.get("education", ""), school=f.get("school", ""),
        major=f.get("major", ""),
        work_years=int(f.get("work_years") or 0),
        expect_salary=int(f.get("expect_salary") or 0),
        current_company=f.get("current_company", ""),
        current_title=f.get("current_title", ""),
        skills_json=skills_json,
        top_school=bool(f.get("top_school")),
        has_management=bool(f.get("has_management")),
        job_changes=int(f.get("job_changes") or 0),
        source_channel=channel,
        file_path=str(dest), file_hash=fhash,
        raw_text=result["raw_text"],
        parse_level="scan" if result["is_scan"] else "ok",
        confidence_json=json.dumps(result["confidence"], ensure_ascii=False),
    )
    db.add(cand)
    db.flush()
    for seq, (st, content) in enumerate(result["sections"], 1):
        db.add(CandidateSection(candidate_id=cand.id, section_type=st,
                                content=content, order_no=seq))
    return "scan" if result["is_scan"] else "ok"


@router.post("/resume/upload")
async def resume_upload(files: list[UploadFile] = File(...),
                        channel: str = Form("其他"),
                        requisition_id: int = Form(0)):
    db = SessionLocal()
    try:
        to_process = []
        for f in files:
            raw = await f.read()
            if not raw:
                continue
            name = f.filename or ""
            if name.lower().endswith(".zip"):
                try:
                    zf = zipfile.ZipFile(io.BytesIO(raw))
                except zipfile.BadZipFile:
                    raise HTTPException(400, "ZIP 文件损坏")
                for zn in zf.namelist():
                    low = zn.lower()
                    if low.endswith(RESUME_EXTS) and not zn.startswith("__MACOSX/"):
                        to_process.append((zn.split("/")[-1], zf.read(zn)))
            else:
                to_process.append((name, raw))

        ok = dup = scan = skip = 0
        for fn, raw in to_process:
            if not fn.lower().endswith(RESUME_EXTS):
                skip += 1
                continue
            r = _save_and_parse(db, fn, raw, channel, requisition_id)
            if r == "dup":
                dup += 1
            elif r == "scan":
                scan += 1
            elif r == "skip":
                skip += 1
            else:
                ok += 1
        _audit(db, "import", "resume", f"新增{ok} 去重{dup} 扫描件{scan} 跳过{skip}")
        db.commit()
        return _redirect(f"/resume?imported=1&ok={ok}&dup={dup}&scan={scan}")
    finally:
        db.close()


# ---------- 简历：初筛 ----------

@router.post("/resume/screen")
def resume_screen(requisition_id: int = Form(0),
                  education: str = Form("不限"), min_years: int = Form(0),
                  city: str = Form("不限"), salary_max: int = Form(0),
                  skills: str = Form(""), skills_mode: str = Form("any"),
                  w_years: int = Form(20), w_school: int = Form(12),
                  w_management: int = Form(14), w_stability: int = Form(12),
                  w_skills: int = Form(20),
                  pass_threshold: int = Form(80), maybe_threshold: int = Form(60)):
    db = SessionLocal()
    try:
        cond = {"education": education, "min_years": min_years, "city": city,
                "salary_max": salary_max, "skills": skills, "skills_mode": skills_mode}
        weights = {"years": w_years, "school": w_school, "management": w_management,
                   "stability": w_stability, "skills": w_skills}
        if requisition_id:
            rq = db.get(Requisition, requisition_id)
            if rq:
                rq.hard_json = json.dumps(cond, ensure_ascii=False)
                rq.soft_json = json.dumps(
                    {"weights": weights, "pass_threshold": pass_threshold,
                     "maybe_threshold": maybe_threshold}, ensure_ascii=False)

        q = db.query(Candidate)
        if requisition_id:
            q = q.filter(Candidate.requisition_id == requisition_id)
        cands = q.all()
        n = 0
        stat = {"pass": 0, "maybe": 0, "reject": 0}
        for c in cands:
            if c.parse_level == "scan" or not c.raw_text:
                continue
            try:
                skills_list = json.loads(c.skills_json or "[]")
            except Exception:
                skills_list = []
            fields = {
                "education": c.education, "work_years": c.work_years,
                "city": c.city, "expect_salary": c.expect_salary,
                "skills": skills_list, "top_school": c.top_school,
                "school": c.school, "has_management": c.has_management,
                "job_changes": c.job_changes,
            }
            r = screen(fields, cond, weights,
                       pass_threshold=pass_threshold,
                       maybe_threshold=maybe_threshold)
            c.bucket = r["bucket"]
            c.score = r["score"]
            c.hard_detail_json = serialize(r["hard_detail"])
            c.score_detail_json = serialize(r["score_detail"])
            c.screened_at = datetime.now()
            if r["bucket"] in stat:
                stat[r["bucket"]] += 1
            n += 1
        _audit(db, "screen", "resume",
               f"岗位id={requisition_id} 筛选{n}份 -> 通过{stat['pass']}/待定{stat['maybe']}/不通过{stat['reject']}；"
               f"条件:学历{education},年限≥{min_years},城市{city},薪资≤{salary_max},技能[{skills}]")
        db.commit()
        return _redirect(f"/resume?rid={requisition_id}")
    finally:
        db.close()


# ---------- 简历：单条操作 ----------

@router.post("/resume/{cid}/delete")
def resume_delete(cid: int):
    db = SessionLocal()
    try:
        c = db.get(Candidate, cid)
        if c:
            _audit(db, "delete", "resume", c.name or f"id={cid}")
            db.query(CandidateSection).filter(CandidateSection.candidate_id == cid).delete()
            db.delete(c)
            db.commit()
        return _redirect("/resume")
    finally:
        db.close()


@router.post("/resume/{cid}/patch")
def resume_patch(cid: int, name: str = Form(""), phone: str = Form(""),
                 email: str = Form(""), city: str = Form(""),
                 education: str = Form(""), school: str = Form(""),
                 work_years: int = Form(0), expect_salary: int = Form(0),
                 current_company: str = Form(""), current_title: str = Form("")):
    """人工修正解析字段（低置信度核对用）"""
    db = SessionLocal()
    try:
        c = db.get(Candidate, cid)
        if not c:
            raise HTTPException(404, "候选人不存在")
        c.name, c.phone, c.email = name.strip(), phone.strip(), email.strip()
        c.city, c.education, c.school = city.strip(), education.strip(), school.strip()
        c.work_years, c.expect_salary = work_years, expect_salary
        c.current_company, c.current_title = current_company.strip(), current_title.strip()
        _audit(db, "patch", "resume", f"id={cid} {c.name}")
        db.commit()
        return _redirect(f"/resume/{cid}")
    finally:
        db.close()


# ---------- 员工档案 ----------

EMPLOYEE_COLS = ["工号", "姓名", "性别", "出生日期", "手机", "邮箱", "部门", "岗位",
                 "入职日期", "试用期到期", "状态", "学历", "身份证", "银行卡", "备注"]


@router.get("/employee/template")
def employee_template():
    import openpyxl
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工档案"
    ws.append(EMPLOYEE_COLS)
    ws.append(["E001", "张三", "男", "1995-03-12", "13800000001", "z@t.com",
               "研发中心", "Java工程师", "2023-09-01", "2023-11-30", "在职", "本科", "", "", ""])
    for col, w in zip("ABCDEFGHIJKLMNO", [8, 10, 6, 12, 14, 18, 12, 14, 12, 12, 8, 8, 20, 20, 24]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_template.xlsx"})


def _row_val(r, idx):
    return str(r[idx] or "").strip() if idx < len(r) and r[idx] is not None else ""


@router.post("/employee/import")
async def employee_import(file: UploadFile = File(...)):
    import openpyxl
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, "Excel 为空")
        header = [str(h or "").strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(header)}
        if "姓名" not in idx:
            raise HTTPException(400, f"缺少列：姓名（表头：{header}）")

        added = 0
        for r in rows[1:]:
            if not r or not any(str(x or "").strip() for x in r):
                continue
            name = _row_val(r, idx.get("姓名", 0))
            if not name:
                continue
            db.add(Employee(
                emp_no=_row_val(r, idx.get("工号", 0)),
                name=name,
                gender=_row_val(r, idx.get("性别", 0)) or "未知",
                birth_date=_parse_date(_row_val(r, idx.get("出生日期", 0))),
                phone=_row_val(r, idx.get("手机", 0)),
                email=_row_val(r, idx.get("邮箱", 0)),
                department=_row_val(r, idx.get("部门", 0)),
                position=_row_val(r, idx.get("岗位", 0)),
                hire_date=_parse_date(_row_val(r, idx.get("入职日期", 0))),
                probation_end=_parse_date(_row_val(r, idx.get("试用期到期", 0))),
                status=_row_val(r, idx.get("状态", 0)) or "在职",
                education=_row_val(r, idx.get("学历", 0)),
                id_card=_row_val(r, idx.get("身份证", 0)),
                bank_card=_row_val(r, idx.get("银行卡", 0)),
                note=_row_val(r, idx.get("备注", 0))))
            added += 1
        _audit(db, "import", "employee", f"文件{file.filename}，新增{added}人")
        db.commit()
        return _redirect("/employees")
    finally:
        db.close()


@router.post("/employee/add")
def employee_add(emp_no: str = Form(""), name: str = Form(...), gender: str = Form("未知"),
                 birth_date: str = Form(""), phone: str = Form(""), email: str = Form(""),
                 department: str = Form(""), position: str = Form(""),
                 hire_date: str = Form(""), probation_end: str = Form(""),
                 status: str = Form("在职"), education: str = Form(""),
                 id_card: str = Form(""), bank_card: str = Form(""), note: str = Form("")):
    db = SessionLocal()
    try:
        db.add(Employee(
            emp_no=emp_no.strip(), name=name.strip(), gender=gender,
            birth_date=_parse_date(birth_date) if birth_date else None,
            phone=phone.strip(), email=email.strip(), department=department.strip(),
            position=position.strip(),
            hire_date=_parse_date(hire_date) if hire_date else None,
            probation_end=_parse_date(probation_end) if probation_end else None,
            status=status, education=education.strip(),
            id_card=id_card.strip(), bank_card=bank_card.strip(), note=note.strip()))
        _audit(db, "add", "employee", name)
        db.commit()
        return _redirect("/employees")
    finally:
        db.close()


@router.post("/employee/{eid}/update")
def employee_update(eid: int, emp_no: str = Form(""), name: str = Form(...),
                    gender: str = Form("未知"), birth_date: str = Form(""),
                    phone: str = Form(""), email: str = Form(""),
                    department: str = Form(""), position: str = Form(""),
                    hire_date: str = Form(""), probation_end: str = Form(""),
                    status: str = Form("在职"), education: str = Form(""),
                    id_card: str = Form(""), bank_card: str = Form(""), note: str = Form("")):
    db = SessionLocal()
    try:
        e = db.get(Employee, eid)
        if not e:
            raise HTTPException(404, "员工不存在")
        e.emp_no, e.name, e.gender = emp_no.strip(), name.strip(), gender
        e.birth_date = _parse_date(birth_date) if birth_date else None
        e.phone, e.email = phone.strip(), email.strip()
        e.department, e.position = department.strip(), position.strip()
        e.hire_date = _parse_date(hire_date) if hire_date else None
        e.probation_end = _parse_date(probation_end) if probation_end else None
        e.status, e.education = status, education.strip()
        # 敏感字段：留空则不修改（编辑页不回填明文，避免泄露；要改则重输完整值）
        if id_card.strip():
            e.id_card = id_card.strip()
        if bank_card.strip():
            e.bank_card = bank_card.strip()
        e.note = note.strip()
        _audit(db, "update", "employee", name)
        db.commit()
        return _redirect(f"/employees/{eid}")
    finally:
        db.close()


@router.post("/employee/{eid}/delete")
def employee_delete(eid: int):
    db = SessionLocal()
    try:
        e = db.get(Employee, eid)
        if e:
            _audit(db, "delete", "employee", e.name)
            db.delete(e)
            db.commit()
        return _redirect("/employees")
    finally:
        db.close()


@router.post("/employee/{eid}/event")
def employee_event(eid: int, event_type: str = Form("其他"),
                   event_date: str = Form(""), detail: str = Form("")):
    db = SessionLocal()
    try:
        e = db.get(Employee, eid)
        if not e:
            raise HTTPException(404, "员工不存在")
        db.add(EmployeeEvent(emp_id=eid, event_type=event_type,
                             event_date=_parse_date(event_date) if event_date else date.today(),
                             detail=detail.strip()))
        _audit(db, "event", "employee", f"{e.name} {event_type} {detail[:50]}")
        db.commit()
        return _redirect(f"/employees/{eid}")
    finally:
        db.close()


@router.post("/employee/event/{evid}/delete")
def employee_event_delete(evid: int):
    db = SessionLocal()
    try:
        ev = db.get(EmployeeEvent, evid)
        if ev:
            db.delete(ev)
            db.commit()
        return _redirect(f"/employees/{ev.emp_id}" if ev else "/employees")
    finally:
        db.close()


# ---------- 考勤异常 ----------

ATT_COLS = ["姓名", "月份", "日期", "异常类型", "原因"]


@router.get("/attendance/template")
def attendance_template():
    import openpyxl
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考勤异常"
    ws.append(ATT_COLS)
    ws.append(["张三", "2026-08", "2026-08-05", "漏打卡", "忘记打卡"])
    ws.append(["李四", "2026-08", "2026-08-07", "迟到", "交通拥堵"])
    for col, w in zip("ABCDE", [10, 10, 14, 10, 30]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_template.xlsx"})


@router.post("/attendance/import")
async def attendance_import(file: UploadFile = File(...)):
    import openpyxl
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, "Excel 为空")
        header = [str(h or "").strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(header)}
        for n in ("姓名", "月份"):
            if n not in idx:
                raise HTTPException(400, f"缺少列：{n}")
        added = 0
        for r in rows[1:]:
            if not r or not any(str(x or "").strip() for x in r):
                continue
            name = _row_val(r, idx["姓名"])
            if not name:
                continue
            db.add(AttendanceException(
                emp_name=name,
                month=_row_val(r, idx["月份"])[:7],
                exception_date=_parse_date(_row_val(r, idx.get("日期", 0))),
                exception_type=_row_val(r, idx.get("异常类型", 0)) or "其他",
                reason=_row_val(r, idx.get("原因", 0)),
                status="待确认"))
            added += 1
        _audit(db, "import", "attendance", f"新增{added}条异常")
        db.commit()
        return _redirect("/attendance")
    finally:
        db.close()


@router.post("/attendance/batch")
def attendance_batch(ids: str = Form(...), status: str = Form("已处理"),
                     handler_note: str = Form("")):
    """批量处理：ids 逗号分隔"""
    db = SessionLocal()
    try:
        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        for iid in id_list:
            a = db.get(AttendanceException, iid)
            if a:
                a.status = status
                if handler_note:
                    a.handler_note = handler_note.strip()
        _audit(db, "batch", "attendance", f"{len(id_list)}条 -> {status}")
        db.commit()
        return _redirect("/attendance")
    finally:
        db.close()


@router.post("/attendance/{aid}/delete")
def attendance_delete(aid: int):
    db = SessionLocal()
    try:
        a = db.get(AttendanceException, aid)
        if a:
            db.delete(a)
            db.commit()
        return _redirect("/attendance")
    finally:
        db.close()


# ---------- 工资明细 ----------

SALARY_COLS = ["月份", "姓名", "基本工资", "绩效", "补贴", "加班费", "社保", "公积金", "个税", "其他扣款", "实发", "备注"]


@router.get("/salary/template")
def salary_template():
    import openpyxl
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工资明细"
    ws.append(SALARY_COLS)
    ws.append(["2026-08", "张三", 15000, 3000, 500, 0, 1600, 1200, 800, 300, 14600, ""])
    for col, w in zip("ABCDEFGHIJKL", [10, 10, 12, 10, 10, 10, 10, 10, 10, 12, 12, 20]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=salary_template.xlsx"})


@router.post("/salary/import")
async def salary_import(file: UploadFile = File(...)):
    import openpyxl
    db = SessionLocal()
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, "Excel 为空")
        header = [str(h or "").strip() for h in rows[0]]
        idx = {h: i for i, h in enumerate(header)}
        for n in ("姓名", "月份"):
            if n not in idx:
                raise HTTPException(400, f"缺少列：{n}")

        def num(k):
            v = _row_val(r, idx.get(k, 0))
            try:
                return int(float(v)) if v else 0
            except ValueError:
                return 0

        added = 0
        for r in rows[1:]:
            if not r or not any(str(x or "").strip() for x in r):
                continue
            name = _row_val(r, idx["姓名"])
            if not name:
                continue
            db.add(SalaryRecord(
                emp_name=name, month=_row_val(r, idx["月份"])[:7],
                base_salary=num("基本工资"), performance=num("绩效"),
                subsidy=num("补贴"), overtime_pay=num("加班费"),
                social_insurance=num("社保"), housing_fund=num("公积金"),
                tax=num("个税"), other_deduct=num("其他扣款"),
                net_salary=num("实发"), note=_row_val(r, idx.get("备注", 0))))
            added += 1
        _audit(db, "import", "salary", f"{added}条明细")
        db.commit()
        return _redirect("/salary")
    finally:
        db.close()


@router.post("/salary/{sid}/delete")
def salary_delete(sid: int):
    db = SessionLocal()
    try:
        s = db.get(SalaryRecord, sid)
        if s:
            db.delete(s)
            db.commit()
        return _redirect("/salary")
    finally:
        db.close()


# ---------- 入离职清单 ----------

ONBOARD_ITEMS = ["身份证复印件收集", "学历学位证书核验", "离职证明收集", "体检报告收集",
                 "银行卡信息登记", "劳动合同签订", "工号开通", "企业微信/邮箱开通",
                 "工牌发放", "门禁权限开通", "办公用品领取", "入职培训完成", "欢迎信发送"]
OFFBOARD_ITEMS = ["离职申请提交", "工作交接完成", "账号权限关闭", "工牌归还",
                  "办公用品归还", "借款/费用结清", "社保停缴", "离职证明开具",
                  "工资结算确认", "员工关怀访谈"]


@router.post("/checklist/start")
def checklist_start(emp_name: str = Form(...), check_type: str = Form("入职"),
                    item_names: str = Form("")):
    """发起入离职清单：item_names 逗号分隔（空则用内置模板）"""
    db = SessionLocal()
    try:
        items = [x.strip() for x in item_names.split(",") if x.strip()] or \
                (ONBOARD_ITEMS if check_type == "入职" else OFFBOARD_ITEMS)
        for it in items:
            db.add(ChecklistItem(emp_name=emp_name.strip(), check_type=check_type,
                                 item_name=it, done=False))
        _audit(db, "add", "checklist", f"{emp_name} {check_type} {len(items)}项")
        db.commit()
        return _redirect("/checklist")
    finally:
        db.close()


@router.post("/checklist/{cid}/toggle")
def checklist_toggle(cid: int):
    db = SessionLocal()
    try:
        c = db.get(ChecklistItem, cid)
        if c:
            c.done = not c.done
            c.done_at = datetime.now() if c.done else None
            _audit(db, "toggle", "checklist",
                   f"{c.emp_name} {c.check_type}:{c.item_name} -> {'完成' if c.done else '未完成'}")
            db.commit()
        return _redirect("/checklist")
    finally:
        db.close()


@router.post("/checklist/{cid}/delete")
def checklist_delete(cid: int):
    db = SessionLocal()
    try:
        c = db.get(ChecklistItem, cid)
        if c:
            db.delete(c)
            db.commit()
        return _redirect("/checklist")
    finally:
        db.close()


# ---------- 文书模板 ----------

@router.post("/doctpl/add")
def doctpl_add(name: str = Form(...), category: str = Form("通知"), content: str = Form(...)):
    db = SessionLocal()
    try:
        db.add(DocTemplate(name=name.strip(), category=category, content=content.strip()))
        _audit(db, "add", "doctpl", name)
        db.commit()
        return _redirect("/docgen")
    finally:
        db.close()


@router.post("/doctpl/{tid}/update")
def doctpl_update(tid: int, name: str = Form(...), category: str = Form("通知"),
                  content: str = Form(...)):
    db = SessionLocal()
    try:
        t = db.get(DocTemplate, tid)
        if t:
            t.name, t.category, t.content = name.strip(), category, content.strip()
            _audit(db, "update", "doctpl", name)
            db.commit()
        return _redirect("/docgen")
    finally:
        db.close()


@router.post("/doctpl/{tid}/delete")
def doctpl_delete(tid: int):
    db = SessionLocal()
    try:
        t = db.get(DocTemplate, tid)
        if t:
            db.delete(t)
            _audit(db, "delete", "doctpl", t.name)
            db.commit()
        return _redirect("/docgen")
    finally:
        db.close()
